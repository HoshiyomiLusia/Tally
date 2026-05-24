"""Import historical bills from the user's Excel collection into Tally.

Usage (inside container):
    python -m scripts.import_excel <username> <root_dir>

Wipes the target user's wallets/transactions/budgets/contacts and re-seeds
defaults, then walks every .xlsx under root_dir, mapping each per-day
per-category cell to one Transaction. Three wallets are created from the
first file's starting balance: 三井住友 (JPY main), NekoPay (JPY e-wallet),
招商银行 (CNY).
"""
import asyncio
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.db import SessionLocal
from app.models import Attachment, Budget, Category, Contact, Currency, Merchant, Transaction, User, Wallet
from app.services.seed import seed_user_defaults


HEADER_MAP_JPY = {
    "餐饮": "餐饮",
    "餐饮/食材": "餐饮",
    "食材": "超市食材",
    "交通": "交通",
    "超市": "超市食材",
    "超市食材": "超市食材",
    "超市，食材": "超市食材",
    "便利店": "零食便利店",
    "药妆": "个护",
    "Amazon": "Amazon",
    "商场": "线下门店",
    "商店购物": "购物",
    "百元店": "生活用品",
    "烘干": "生活用品",
    "网购": "淘宝京东",
    "商场日用品": "线下门店",
    "商场，日用品": "线下门店",
    "娱乐": "娱乐",
    "其他": "其他",
    "手续费": "其他",
    "月度": "订阅",
    "换汇": "其他收入",
    "换汇收入": "其他收入",
    "换汇，收入": "其他收入",
    "收入": "工资",
    "报销": "报销",
    "存入": "其他收入",
}

HEADER_MAP_CNY = {
    "娱乐(R)": "娱乐",
    "娱乐（R）": "娱乐",
    "其他(R)": "其他",
    "其他（R）": "其他",
    "月度(R)": "订阅",
    "月度（R）": "订阅",
    "手续费(R)": "其他",
    "手续费（R）": "其他",
    "换汇(R)": "其他收入",
    "换汇（R）": "其他收入",
    "换汇收入(R)": "其他收入",
    "换汇收入（R）": "其他收入",
    "换汇，收入(R)": "其他收入",
    "换汇，收入（R）": "其他收入",
    "报销(R)": "报销",
    "报销（R）": "报销",
    "收入(R)": "工资",
    "收入（R）": "工资",
}

INCOME_CATS = {"工资", "生活费", "报销", "退款", "投资收益", "其他收入"}


def normalize(h):
    if h is None:
        return None
    s = str(h).strip().replace(" ", "").replace("　", "")
    return s


def parse_year_month(sheet_name, filename):
    m = re.match(r"^(\d{4})\.(\d{1,2})", sheet_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d{4})\.(\d{1,2})(?!\d)", filename)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def detect_sides(headers):
    """Find column ranges for JPY / CNY / totals.

    Returns (jpy_balance_col, cny_balance_col). Cols <= jpy_balance_col are
    JPY side; cols in (jpy_balance_col, cny_balance_col] are CNY side; cols
    beyond cny_balance_col are totals/aggregates and skipped.
    """
    jpy_bal = None
    cny_bal = None
    for c in sorted(headers):
        h = headers[c]
        if h == "余额" and jpy_bal is None:
            jpy_bal = c
        elif h in ("余额(R)", "余额（R）") and cny_bal is None:
            cny_bal = c
    return jpy_bal, cny_bal


def map_header(header, side):
    """Return (tally_category_name, currency) or None to skip."""
    if side == "SKIP" or not header:
        return None
    if side == "JPY":
        if header in HEADER_MAP_JPY:
            return HEADER_MAP_JPY[header], "JPY"
        return None
    # CNY side: try R-suffixed map first, then strip suffix and try JPY map
    if header in HEADER_MAP_CNY:
        return HEADER_MAP_CNY[header], "CNY"
    base = header.replace("(R)", "").replace("（R）", "").strip()
    if base in HEADER_MAP_JPY:
        return HEADER_MAP_JPY[base], "CNY"
    return None


def col_side(c, jpy_bal, cny_bal):
    if jpy_bal is None:
        return "JPY"
    if c <= jpy_bal:
        return "JPY"
    if cny_bal is not None and c <= cny_bal:
        return "CNY"
    return "SKIP"


async def import_for_user(username, root_dir):
    async with SessionLocal() as session:
        user = (await session.execute(select(User).where(User.username == username))).scalar_one_or_none()
        if not user:
            print(f"user {username!r} not found")
            sys.exit(1)

        await session.execute(delete(Attachment).where(Attachment.user_id == user.id))
        await session.execute(delete(Transaction).where(Transaction.user_id == user.id))
        await session.execute(delete(Budget).where(Budget.user_id == user.id))
        await session.execute(delete(Wallet).where(Wallet.user_id == user.id))
        await session.execute(delete(Merchant).where(Merchant.user_id == user.id))
        await session.execute(delete(Contact).where(Contact.user_id == user.id))
        await session.execute(delete(Category).where(Category.user_id == user.id))
        await session.commit()
        await seed_user_defaults(session, user.id)

        digits_map = {c: d for c, d in (await session.execute(select(Currency.code, Currency.decimal_digits))).all()}
        cat_lookup = {c.name: c.id for c in (await session.execute(select(Category).where(Category.user_id == user.id))).scalars().all()}

        files = sorted(p for p in Path(root_dir).rglob("*.xlsx") if "模板" not in p.name)
        if not files:
            print("no xlsx files found")
            return

        # Pull starting balances from the chronologically earliest *sheet*
        # across all files. Alphabetic sort puts "10" before "4", so we
        # have to look at year+month from sheet names.
        first_main_jpy = 0
        first_main_cny = 0
        first_nekopay = 0
        earliest_main = (9999, 12)
        earliest_neko = (9999, 12)
        for f in files:
            wb = load_workbook(f, data_only=True)
            for sn in wb.sheetnames:
                year, month = parse_year_month(sn, f.name)
                if not year:
                    continue
                s = wb[sn]
                headers = {c: normalize(s.cell(3, c).value) for c in range(1, s.max_column + 1)}
                is_neko = "NekoPay" in sn
                key = (year, month)
                target_earliest = earliest_neko if is_neko else earliest_main
                if key >= target_earliest:
                    continue
                for c, h in headers.items():
                    if h == "余额":
                        v = s.cell(4, c).value
                        if isinstance(v, (int, float)):
                            if is_neko:
                                first_nekopay = int(round(v))
                                earliest_neko = key
                            else:
                                first_main_jpy = int(round(v))
                                earliest_main = key
                    elif h in ("余额(R)", "余额（R）") and not is_neko:
                        v = s.cell(4, c).value
                        if isinstance(v, (int, float)):
                            first_main_cny = int(round(v * 100))
        print(f"initial balances: SMBC={first_main_jpy} JPY (from {earliest_main}), "
              f"NekoPay={first_nekopay} JPY (from {earliest_neko}), CMB={first_main_cny} fen")

        smbc = Wallet(user_id=user.id, name="三井住友銀行", type="bank", currency_code="JPY",
                       initial_balance=first_main_jpy, color="#0f7d3a")
        nekopay = Wallet(user_id=user.id, name="NekoPay", type="e_wallet", currency_code="JPY",
                          initial_balance=first_nekopay, color="#0b59a8")
        cmb = Wallet(user_id=user.id, name="招商银行", type="bank", currency_code="CNY",
                      initial_balance=first_main_cny, color="#c8102e")
        session.add_all([smbc, nekopay, cmb])
        await session.commit()
        for w in (smbc, nekopay, cmb):
            await session.refresh(w)

        skipped_headers = set()
        total = 0

        for f in files:
            wb = load_workbook(f, data_only=True)
            file_total = 0
            for sn in wb.sheetnames:
                year, month = parse_year_month(sn, f.name)
                if not year:
                    continue
                is_neko = "NekoPay" in sn
                s = wb[sn]
                headers = {}
                for c in range(1, s.max_column + 1):
                    h = normalize(s.cell(3, c).value)
                    if h:
                        headers[c] = h
                jpy_bal, cny_bal = detect_sides(headers)

                for r in range(5, s.max_row + 1):
                    day = s.cell(r, 2).value
                    if not isinstance(day, int) or day < 1 or day > 31:
                        continue
                    try:
                        tx_date = date(year, month, day)
                    except ValueError:
                        continue

                    for col, hdr in headers.items():
                        side = col_side(col, jpy_bal, cny_bal)
                        mapping = map_header(hdr, side)
                        if mapping is None:
                            if side != "SKIP":
                                skipped_headers.add(f"{hdr}[{side}]")
                            continue
                        cat_name, currency = mapping
                        v = s.cell(r, col).value
                        if v is None or not isinstance(v, (int, float)) or v == 0:
                            continue
                        cat_id = cat_lookup.get(cat_name)
                        if cat_id is None:
                            continue
                        # Sign determines kind. The categories named like
                        # "其他收入" come from 换汇/换汇，收入 columns where
                        # negative values are JPY-out exchanges (expense).
                        kind = "income" if v > 0 else "expense"
                        if currency == "CNY":
                            wallet = cmb
                        else:
                            wallet = nekopay if is_neko else smbc
                        if wallet.currency_code != currency:
                            continue
                        d = digits_map.get(currency, 2)
                        amount = int(round(abs(float(v)) * (10 ** d)))
                        if amount == 0:
                            continue
                        session.add(Transaction(
                            user_id=user.id,
                            wallet_id=wallet.id,
                            category_id=cat_id,
                            amount=amount,
                            currency_code=currency,
                            kind=kind,
                            occurred_on=tx_date,
                            note=f"{hdr} [{sn}]",
                        ))
                        file_total += 1
                        total += 1
                await session.commit()
            print(f"  {f.name}: +{file_total}")

        print(f"\nimported {total} transactions for {username!r}")
        if skipped_headers:
            print(f"unmapped headers skipped (informational): {sorted(skipped_headers)}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("usage: python -m scripts.import_excel <username> <root_dir>")
        sys.exit(1)
    asyncio.run(import_for_user(sys.argv[1], sys.argv[2]))
